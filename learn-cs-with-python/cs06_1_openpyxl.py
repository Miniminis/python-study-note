"""
*120분*

절차지향형 vs 객체지향 프로그래밍
"""

"""
절차지향형 프로그래밍
= 프로시저 = 함수 
한눈에 이 프로그램이 어떤일을 하는지 알 수 있게 프로그래밍하는 것
함수들의 이름만봐도 이 프로그램의 용도를 알 수 있다. 
"""

"""
pip install openpyxl 을 통해서 엑셀 파일을 읽어들일 수 있도록 한다.
"""


from openpyxl import load_workbook
wb = load_workbook('cs06_exam.xlsx')  # load_workbook 이용해서 엑셀 파일 임포트 
print(wb.sheetnames)

ws = wb.active  # 활성화되어있는 시트 가져오기
print(ws)

g = ws.rows  # 모든 행을 가져와 객체로 반환
cells = next(g)
print(cells)

keys = []
for cell in cells:
    keys.append(cell.value)    # cell 의 value 에 접근하여 가져옴 

print(keys)


student_data = []
for row in g:
    dic = {k: c.value for k, c in zip(keys, row)}  # 딕셔너리 컴프리헨션
    student_data.append(dic)

print(student_data)

